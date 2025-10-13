"""
EXCEL WITH EMBEDDED LOGOS
Creates Excel file with actual images in cells - can upload to Google Sheets!

REQUIREMENTS: pip install openpyxl pillow

Run this AFTER downloading logos
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from datetime import datetime

BATCH_NUMBER = '54'
LOGOS_FOLDER = f'batch_{BATCH_NUMBER}_logos'

def create_excel_with_images():
    print("=" * 70)
    print(f"📊 CREATING EXCEL WITH EMBEDDED IMAGES - BATCH {BATCH_NUMBER}")
    print("=" * 70)
    
    logos_path = Path(LOGOS_FOLDER)
    
    if not logos_path.exists():
        print(f"\n❌ ERROR: Folder '{LOGOS_FOLDER}' not found!")
        return
    
    # Get all image files (including converted PNGs from SVG)
    image_files = []
    for ext in ['*.png', '*.jpg', '*.jpeg', '*.gif']:
        image_files.extend(logos_path.glob(ext))
    
    # Also check for SVG files (we'll note them but prefer PNG versions)
    svg_files = list(logos_path.glob('*.svg'))
    
    if not image_files:
        print(f"\n❌ No compatible images found in {LOGOS_FOLDER}/")
        if svg_files:
            print(f"   Found {len(svg_files)} SVG files but they need conversion to PNG")
            print(f"   Run the download script again with cairosvg installed")
        return
    
    print(f"\n📸 Found {len(image_files)} images")
    
    # Group by brand
    brands = {}
    for img_path in image_files:
        filename = img_path.stem
        parts = filename.rsplit('_logo', 1)
        brand_name = parts[0].replace('_', ' ')
        
        if brand_name not in brands:
            brands[brand_name] = {'logo1': None, 'logo2': None, 'logo3': None}
        
        if 'logo1' in filename.lower():
            brands[brand_name]['logo1'] = img_path
        elif 'logo2' in filename.lower():
            brands[brand_name]['logo2'] = img_path
        elif 'logo3' in filename.lower():
            brands[brand_name]['logo3'] = img_path
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f'Batch {BATCH_NUMBER} Logos'
    
    # Set column widths and row heights
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    
    # Headers
    ws['A1'] = 'Brand'
    ws['B1'] = 'Logo 1'
    ws['C1'] = 'Logo 2'
    ws['D1'] = 'Logo 3'
    
    # Style headers
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = ws[cell].font.copy(bold=True)
    
    ws.row_dimensions[1].height = 20
    
    print("\n🖼️ Embedding images into Excel...")
    
    # Add brands and images
    row = 2
    for brand_name, logos in sorted(brands.items()):
        print(f"   {brand_name}")
        
        ws[f'A{row}'] = brand_name
        
        # Set row height for images
        ws.row_dimensions[row].height = 80
        
        # Add each logo
        for idx, logo_key in enumerate(['logo1', 'logo2', 'logo3'], start=2):
            logo_path = logos[logo_key]
            
            if logo_path and logo_path.exists():
                try:
                    # Create image object
                    img = XLImage(str(logo_path))
                    
                    # Resize to fit cell (max 150px width, 100px height)
                    max_width = 150
                    max_height = 100
                    
                    # Calculate scaling
                    width_scale = max_width / img.width if img.width > max_width else 1
                    height_scale = max_height / img.height if img.height > max_height else 1
                    scale = min(width_scale, height_scale)
                    
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)
                    
                    # Position in cell
                    col_letter = get_column_letter(idx)
                    cell_ref = f'{col_letter}{row}'
                    
                    # Add image to cell
                    img.anchor = cell_ref
                    ws.add_image(img)
                    
                except Exception as e:
                    ws[f'{get_column_letter(idx)}{row}'] = f'Error: {str(e)[:20]}'
        
        row += 1
    
    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = f'batch_{BATCH_NUMBER}_with_images_{timestamp}.xlsx'
    wb.save(output_file)
    
    print(f"\n✅ Excel file created with embedded images!")
    print(f"   📄 {output_file}")
    print(f"\n🚀 TO USE:")
    print(f"   1. Open the file in Excel/LibreOffice")
    print(f"   2. Upload to Google Drive")
    print(f"   3. Open with Google Sheets")
    print(f"   4. Images will be preserved!")
    print(f"\n💡 OR:")
    print(f"   - Copy cells directly from Excel")
    print(f"   - Paste into Google Sheets")
    print(f"   - Images copy over!")
    print("=" * 70)

if __name__ == '__main__':
    create_excel_with_images()